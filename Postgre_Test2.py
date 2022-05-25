
import psycopg2
import openpyxl

from tkinter import *

# DB 커넥션 정보 세팅
conn_string = "host='localhost' dbname = 'postgres' user = 'postgres' password = 'inzent01'"
# DB 커넥션 
conn = psycopg2.connect(conn_string)
cur = conn.cursor()

#cur.execute("INSERT INTO films (code,title,did,date_prod,kind,len) VALUES (%s,%s,%s,%s,%s,%s)",("0004","test4",3,"2022-05-25","wwww","0"))
# 일반적인 쿼리
cur.execute("SELECT * FROM films;")
result = cur.fetchall()
print ( result )

# 엑셀파일 열기
wb = openpyxl.load_workbook('films.xlsx')
 
# 현재 Active Sheet 얻기
ws = wb.active


row_index = 0

for row in result:
    row_index = row_index + 1   # 행 인덱스
    
    print("Data row = (%s, %s, %s, %s, %s, %s)" %( str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]) ) )
    ws.cell(row = row_index, column=1).value = row[0]
    ws.cell(row = row_index, column=2).value = row[1]
    ws.cell(row = row_index, column=3).value = row[2]
    ws.cell(row = row_index, column=4).value = row[3]
    ws.cell(row = row_index, column=5).value = row[4]
    ws.cell(row = row_index, column=6).value = row[5]				

# 데이터를 변경했다면 반드시 .commit() 해주어야 한다
conn.commit()

# 커서를 닫고 연걸을 종료한다.
cur.close()
conn.close()

# 엑셀 파일 저장
wb.save("films1.xlsx")
wb.close()

